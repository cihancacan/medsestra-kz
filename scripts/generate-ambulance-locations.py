#!/usr/bin/env python3
"""Generate the ambulance SEO location catalogue from Kazakhstan's KATO file.

Usage:
    python scripts/generate-ambulance-locations.py /path/to/KATO.xlsx

The generated TypeScript contains every official city and every official
intra-city district present in KATO. It intentionally excludes villages and
rural districts: publishing a medical landing page for every settlement would
create low-value doorway pages rather than a useful national directory.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


TRANSLIT = {
    "а": "a", "ә": "a", "б": "b", "в": "v", "г": "g", "ғ": "g",
    "д": "d", "е": "e", "ё": "yo", "ж": "zh", "з": "z", "и": "i",
    "й": "y", "к": "k", "қ": "q", "л": "l", "м": "m", "н": "n",
    "ң": "ng", "о": "o", "ө": "o", "п": "p", "р": "r", "с": "s",
    "т": "t", "у": "u", "ұ": "u", "ү": "u", "ф": "f", "х": "kh",
    "һ": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "shch", "ъ": "",
    "ы": "y", "і": "i", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}

ENGLISH_CITY_NAMES = {
    "Семей": "Semey", "Курчатов": "Kurchatov", "Аягоз": "Ayagoz",
    "Шар": "Shar", "Кокшетау": "Kokshetau", "Косшы": "Kosshy",
    "Степногорск": "Stepnogorsk", "Акколь": "Akkol", "Атбасар": "Atbasar",
    "Макинск": "Makinsk", "Степняк": "Stepnyak", "Ерейментау": "Ereymentau",
    "Есиль": "Esil", "Державинск": "Derzhavinsk", "Щучинск": "Shchuchinsk",
    "Актобе": "Aktobe", "Алга": "Alga", "Кандыагаш": "Kandyagash",
    "Эмба": "Emba", "Жем": "Zhem", "Темир": "Temir", "Хромтау": "Khromtau",
    "Шалкар": "Shalkar", "Қонаев": "Konaev", "Алатау": "Alatau",
    "Есик": "Esik", "Каскелен": "Kaskelen", "Талгар": "Talgar",
    "Атырау": "Atyrau", "Кульсары": "Kulsary", "Уральск": "Oral",
    "Аксай": "Aksai", "Тараз": "Taraz", "Жанатас": "Zhanatas",
    "Каратау": "Karatau", "Шу": "Shu", "Талдыкорган": "Taldykorgan",
    "Текели": "Tekeli", "Ушарал": "Usharal", "Уштобе": "Ushtobe",
    "Жаркент": "Zharkent", "Саркан": "Sarkand", "Караганда": "Karaganda",
    "Балхаш": "Balkhash", "Приозерск": "Priozersk", "Сарань": "Saran",
    "Темиртау": "Temirtau", "Шахтинск": "Shakhtinsk", "Абай": "Abai",
    "Каркаралинск": "Karkaraly", "Костанай": "Kostanay", "Аркалык": "Arkalyk",
    "Лисаковск": "Lisakovsk", "Рудный": "Rudny", "Житикара": "Zhitikara",
    "Тобыл": "Tobyl", "Кызылорда": "Kyzylorda", "Байконыр": "Baikonur",
    "Аральск": "Aralsk", "Казалинск": "Kazalinsk", "Актау": "Aktau",
    "Жанаозен": "Zhanaozen", "Форт-Шевченко": "Fort-Shevchenko",
    "Павлодар": "Pavlodar", "Аксу": "Aksu", "Экибастуз": "Ekibastuz",
    "Петропавловск": "Petropavl", "Булаево": "Bulaevo",
    "Мамлютка": "Mamlyutka", "Сергеевка": "Sergeevka",
    "Тайынша": "Taiynsha", "Туркестан": "Turkistan", "Арысь": "Arys",
    "Кентау": "Kentau", "Жетысай": "Zhetysai", "Сарыагаш": "Saryagash",
    "Ленгер": "Lenger", "Шардара": "Shardara", "Жезказган": "Zhezkazgan",
    "Каражал": "Karazhal", "Сатпаев": "Satpayev",
    "Усть-Каменогорск": "Oskemen", "Риддер": "Ridder", "Зайсан": "Zaysan",
    "Алтай": "Altai", "Серебрянск": "Serebryansk",
    "Шемонаиха": "Shemonaikha", "Астана": "Astana", "Алматы": "Almaty",
    "Шымкент": "Shymkent",
}

ENGLISH_DISTRICT_NAMES = {
    "151011100": "Astana District",
    "151013100": "Almaty District",
    "311011100": "Aulieata District",
    "311013100": "Zhibek Zholy District",
    "351011100": "Alikhan Bokeikhan District",
    "351013100": "Kazybek Bi District",
    "711110000": "Almaty District",
    "711210000": "Esil District",
    "711310000": "Saryarka District",
    "711410000": "Baikonur District",
    "711510000": "Nura District",
    "711610000": "Sarayshyk District",
    "751110000": "Almaly District",
    "751210000": "Alatau District",
    "751310000": "Auezov District",
    "751410000": "Bostandyk District",
    "751510000": "Zhetysu District",
    "751710000": "Medeu District",
    "751810000": "Nauryzbay District",
    "751910000": "Turksib District",
    "791110000": "Abai District",
    "791310000": "Al-Farabi District",
    "791510000": "Enbekshi District",
    "791710000": "Karatau District",
    "791910000": "Turan District",
}

CITY_CODE_FOR_DISTRICT_PREFIX = {
    "1510": "151010000",
    "3110": "311010000",
    "3510": "351010000",
    "711": "710000000",
    "751": "750000000",
    "791": "790000000",
}


def clean_name(value: str, prefixes: tuple[str, ...]) -> str:
    value = value.strip()
    for prefix in prefixes:
        if value.startswith(prefix):
            return value[len(prefix):].strip()
    return value


def transliterate(value: str) -> str:
    chars: list[str] = []
    for char in value.lower():
        chars.append(TRANSLIT.get(char, char))
    value = "".join(chars)
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def title_from_slug(slug: str) -> str:
    return " ".join(part.capitalize() for part in slug.split("-"))


def parent_city_code(code: str) -> str | None:
    for prefix, city_code in CITY_CODE_FOR_DISTRICT_PREFIX.items():
        if code.startswith(prefix):
            return city_code
    return None


def main() -> None:
    if len(sys.argv) != 2:
        raise SystemExit("Usage: generate-ambulance-locations.py /path/to/KATO.xlsx")

    source = Path(sys.argv[1])
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    regions: dict[str, tuple[str, str]] = {}
    for row in rows:
        code = str(row[0])
        if code.endswith("0000000"):
            regions[code[:2]] = (str(row[7]), str(row[6]))

    cities: list[dict[str, object]] = []
    city_by_code: dict[str, dict[str, object]] = {}
    for row in rows:
        code = str(row[0])
        ru_raw = str(row[7] or "")
        if not re.match(r"^г\.(?!а\.)", ru_raw):
            continue

        ru = clean_name(ru_raw, ("г.",))
        kk = clean_name(str(row[6] or ""), ("қ.", "ќ."))
        en = ENGLISH_CITY_NAMES.get(ru, title_from_slug(transliterate(ru)))
        region_ru, region_kk = regions.get(code[:2], ("Казахстан", "Қазақстан"))
        item: dict[str, object] = {
            "id": code,
            "type": "city",
            "names": {"ru": ru, "kk": kk, "en": en},
            "slugs": {
                "ru": transliterate(en),
                "kk": transliterate(en),
                "en": transliterate(en),
            },
            "region": {"ru": region_ru, "kk": region_kk, "en": "Kazakhstan"},
        }
        cities.append(item)
        city_by_code[code] = item

    districts: list[dict[str, object]] = []
    for row in rows:
        code = str(row[0])
        ru_raw = str(row[7] or "")
        if "район" not in ru_raw.lower() or str(row[5]) != "1":
            continue

        city_code = parent_city_code(code)
        parent = city_by_code.get(city_code or "")
        if not parent:
            continue

        kk = str(row[6]).strip()
        ru = ru_raw.strip()
        en = ENGLISH_DISTRICT_NAMES.get(code, title_from_slug(transliterate(kk)) + " District")
        district_base_en = re.sub(r"\s+District$", "", en)
        districts.append({
            "id": code,
            "type": "district",
            "parentId": city_code,
            "names": {"ru": ru, "kk": kk, "en": en},
            "slugs": {
                "ru": transliterate(ru),
                "kk": transliterate(kk),
                "en": transliterate(district_base_en) + "-district",
            },
            "region": parent["region"],
        })

    locations = sorted(
        [*cities, *districts],
        key=lambda item: (
            str(item.get("parentId") or item["id"]),
            0 if item["type"] == "city" else 1,
            str(item["id"]),
        ),
    )

    output = Path("lib/ambulance-locations.generated.ts")
    payload = json.dumps(locations, ensure_ascii=False, indent=2)
    output.write_text(
        "/* Generated from KATO NK RK 11-2025, updated 17 July 2026. */\n"
        "/* Source: https://stat.gov.kz/ru/classifiers/statistical/21/ */\n"
        "import type { AmbulanceLocation } from '@/lib/ambulance-seo-data';\n\n"
        f"export const ambulanceLocations: AmbulanceLocation[] = {payload};\n",
        encoding="utf-8",
    )
    print(f"Generated {len(cities)} cities and {len(districts)} city districts in {output}")


if __name__ == "__main__":
    main()
